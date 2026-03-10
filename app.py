from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import sqlite3
import os
import io
import base64
import pickle
from datetime import datetime, timedelta
from functools import wraps
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production-render")

# ── CONFIG FROM ENV VARS (set on Render dashboard) ───────────────────────────
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
SMTP_HOST      = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT      = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER      = os.environ.get("SMTP_USER", "")
SMTP_PASS      = os.environ.get("SMTP_PASS", "")
ALERT_FROM     = os.environ.get("ALERT_FROM", SMTP_USER)
ALERT_TO       = os.environ.get("ALERT_TO", "")

# ── PATHS ─────────────────────────────────────────────────────────────────────
DB_PATH = os.environ.get("DB_PATH", "database/students.db")
os.makedirs("database", exist_ok=True)
os.makedirs("models",   exist_ok=True)
os.makedirs("images",   exist_ok=True)

# ── DATABASE ──────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS students(
        usn TEXT PRIMARY KEY, name TEXT, email TEXT, registered_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS attendance(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usn TEXT, name TEXT, date TEXT, time TEXT)""")
    conn.commit()
    conn.close()

init_db()

# ── AUTH DECORATOR ────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

# ── PAGE ROUTES ───────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET"])
def login_page():
    if session.get("logged_in"):
        return redirect(url_for("dashboard"))
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def do_login():
    data = request.json
    if data.get("username") == ADMIN_USERNAME and data.get("password") == ADMIN_PASSWORD:
        session["logged_in"] = True
        session["username"]  = data.get("username")
        return jsonify({"success": True})
    return jsonify({"success": False, "error": "Invalid username or password"}), 401

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/")
@login_required
def dashboard():
    return render_template("index.html", username=session.get("username", "Admin"))

# ── API: STATS ────────────────────────────────────────────────────────────────
@app.route("/api/stats")
@api_login_required
def stats():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) as n FROM students")
    total   = c.fetchone()["n"]
    today   = datetime.now().strftime("%Y-%m-%d")
    c.execute("SELECT COUNT(DISTINCT usn) as n FROM attendance WHERE date=?", (today,))
    present = c.fetchone()["n"]

    days_data = []
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        c.execute("SELECT COUNT(DISTINCT usn) as n FROM attendance WHERE date=?", (d,))
        days_data.append({"date": d, "count": c.fetchone()["n"]})

    monthly_data = []
    for i in range(5, -1, -1):
        m = (datetime.now().replace(day=1) - timedelta(days=i*28)).strftime("%Y-%m")
        c.execute("SELECT COUNT(*) as n FROM attendance WHERE date LIKE ?", (f"{m}%",))
        monthly_data.append({"month": m, "count": c.fetchone()["n"]})

    conn.close()
    return jsonify({"total_students": total, "present_today": present,
                    "absent_today": total - present,
                    "weekly": days_data, "monthly": monthly_data})

# ── API: STUDENTS ─────────────────────────────────────────────────────────────
@app.route("/api/students")
@api_login_required
def get_students():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT usn, name, email, registered_at FROM students ORDER BY name")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/register", methods=["POST"])
@api_login_required
def register_student():
    data  = request.json
    usn   = data.get("usn",   "").strip()
    name  = data.get("name",  "").strip()
    email = data.get("email", "").strip()
    img   = data.get("image", "")
    if not usn or not name:
        return jsonify({"success": False, "error": "USN and Name are required"}), 400
    try:
        if img and img.startswith("data:image"):
            _, enc = img.split(",", 1)
            with open(f"images/{usn}.jpg", "wb") as f:
                f.write(base64.b64decode(enc))
        conn = get_db(); c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO students VALUES (?,?,?,?)",
                  (usn, name, email, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit(); conn.close()
        return jsonify({"success": True, "message": f"{name} registered!"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/delete_student/<usn>", methods=["DELETE"])
@api_login_required
def delete_student(usn):
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM students WHERE usn=?", (usn,))
    conn.commit(); conn.close()
    for ext in [".jpg", ".jpeg", ".png"]:
        p = f"images/{usn}{ext}"
        if os.path.exists(p): os.remove(p)
    return jsonify({"success": True})

# ── API: ENCODE FACES ─────────────────────────────────────────────────────────
@app.route("/api/encode", methods=["POST"])
@api_login_required
def encode_faces():
    try:
        import face_recognition
        encodings, names = [], []
        files = [f for f in os.listdir("images") if f.lower().endswith(('.jpg','.jpeg','.png'))]
        if not files:
            return jsonify({"success": False, "error": "No images found. Register students first."})
        for file in files:
            img = face_recognition.load_image_file(f"images/{file}")
            enc = face_recognition.face_encodings(img)
            if enc:
                encodings.append(enc[0])
                names.append(file.rsplit(".", 1)[0])
        with open("models/encodings.pkl", "wb") as f:
            pickle.dump({"encodings": encodings, "names": names}, f)
        return jsonify({"success": True, "message": f"Encoded {len(encodings)} face(s)!"})
    except ImportError:
        return jsonify({"success": False, "error": "face_recognition not available on this server."})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ── API: RECOGNIZE FACE ───────────────────────────────────────────────────────
@app.route("/api/recognize_face", methods=["POST"])
@api_login_required
def recognize_face():
    data     = request.json
    img_data = data.get("image", "")
    if not img_data:
        return jsonify({"recognized": False, "face_detected": False})
    try:
        import face_recognition
        import numpy as np
        from PIL import Image

        _, enc = img_data.split(",", 1)
        img    = Image.open(io.BytesIO(base64.b64decode(enc))).convert("RGB")
        arr    = np.array(img)
        locs   = face_recognition.face_locations(arr)
        if not locs:
            return jsonify({"recognized": False, "face_detected": False})

        encs = face_recognition.face_encodings(arr, locs)
        pkl  = "models/encodings.pkl"
        if not os.path.exists(pkl):
            return jsonify({"recognized": False, "face_detected": True,
                            "error": "Run Encode Faces first."})
        with open(pkl, "rb") as f:
            known = pickle.load(f)

        for encoding, loc in zip(encs, locs):
            matches = face_recognition.compare_faces(known["encodings"], encoding, tolerance=0.5)
            if True in matches:
                usn  = known["names"][matches.index(True)]
                conn = get_db(); c = conn.cursor()
                c.execute("SELECT name FROM students WHERE usn=?", (usn,))
                row  = c.fetchone(); conn.close()
                name = row["name"] if row else usn
                return jsonify({"recognized": True, "face_detected": True,
                                "usn": usn, "name": name, "box": list(loc)})
        return jsonify({"recognized": False, "face_detected": True})
    except ImportError:
        return jsonify({"recognized": False, "face_detected": False,
                        "error": "face_recognition not installed."})
    except Exception as e:
        return jsonify({"recognized": False, "face_detected": False, "error": str(e)})

# ── API: MARK ATTENDANCE ──────────────────────────────────────────────────────
@app.route("/api/mark_attendance", methods=["POST"])
@api_login_required
def mark_attendance():
    data = request.json
    usn  = data.get("usn", "").strip()
    name = data.get("name", "")
    if not usn:
        return jsonify({"success": False, "error": "USN required"}), 400
    conn  = get_db(); c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("SELECT id FROM attendance WHERE usn=? AND date=?", (usn, today))
    if c.fetchone():
        conn.close()
        return jsonify({"success": False, "already": True, "error": "Already marked today"})
    if not name:
        c.execute("SELECT name FROM students WHERE usn=?", (usn,))
        row  = c.fetchone(); name = row["name"] if row else "Unknown"
    now = datetime.now()
    c.execute("INSERT INTO attendance (usn,name,date,time) VALUES (?,?,?,?)",
              (usn, name, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")))
    conn.commit(); conn.close()
    return jsonify({"success": True, "message": f"Attendance marked for {name}"})

# ── API: GET ATTENDANCE ───────────────────────────────────────────────────────
@app.route("/api/attendance")
@api_login_required
def get_attendance():
    date = request.args.get("date", "")
    usn  = request.args.get("usn",  "")
    conn = get_db(); c = conn.cursor()
    q = "SELECT * FROM attendance WHERE 1=1"; p = []
    if date: q += " AND date=?"; p.append(date)
    if usn:  q += " AND usn=?";  p.append(usn)
    q += " ORDER BY date DESC, time DESC"
    c.execute(q, p)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(rows)

# ── API: ANALYTICS ────────────────────────────────────────────────────────────
@app.route("/api/analytics")
@api_login_required
def analytics():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) as n FROM students")
    total = c.fetchone()["n"]
    c.execute("SELECT DISTINCT date FROM attendance")
    total_classes = len(c.fetchall())
    c.execute("""SELECT s.usn, s.name, COUNT(a.id) as attended
                 FROM students s LEFT JOIN attendance a ON s.usn=a.usn
                 GROUP BY s.usn ORDER BY attended DESC""")
    rows = c.fetchall(); conn.close()
    data = []
    for r in rows:
        pct = round((r["attended"] / total_classes) * 100, 2) if total_classes else 0
        data.append({"usn": r["usn"], "name": r["name"], "attended": r["attended"],
                     "total": total_classes, "percentage": pct,
                     "status": "Good" if pct >= 75 else ("Average" if pct >= 50 else "Low")})
    return jsonify({"total_students": total, "total_classes": total_classes, "data": data})

# ── API: EXPORT CSV / EXCEL ───────────────────────────────────────────────────
@app.route("/api/export")
@api_login_required
def export_attendance():
    fmt  = request.args.get("format", "csv")
    date = request.args.get("date", "")
    usn  = request.args.get("usn",  "")
    conn = get_db(); c = conn.cursor()
    q = """SELECT a.usn, a.name, a.date, a.time, s.email
           FROM attendance a LEFT JOIN students s ON a.usn=s.usn WHERE 1=1"""
    p = []
    if date: q += " AND a.date=?"; p.append(date)
    if usn:  q += " AND a.usn=?";  p.append(usn)
    q += " ORDER BY a.date DESC, a.time DESC"
    c.execute(q, p); rows = c.fetchall(); conn.close()

    if fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Attendance"
            headers = ["USN", "Name", "Email", "Date", "Time"]
            hfill   = PatternFill("solid", fgColor="0D1117")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = Font(bold=True, color="00E5FF")
                cell.fill      = hfill
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 28
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 12
            for row in rows:
                ws.append([row["usn"], row["name"], row["email"] or "", row["date"], row["time"]])
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return send_file(buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"attendance_{date or 'all'}.xlsx")
        except ImportError:
            return jsonify({"error": "openpyxl not installed"}), 500
    else:
        lines = ["USN,Name,Email,Date,Time"]
        for r in rows:
            lines.append(f'{r["usn"]},{r["name"]},{r["email"] or ""},{r["date"]},{r["time"]}')
        buf = io.BytesIO("\n".join(lines).encode())
        return send_file(buf, mimetype="text/csv", as_attachment=True,
                         download_name=f"attendance_{date or 'all'}.csv")

# ── API: EMAIL ALERTS ─────────────────────────────────────────────────────────
@app.route("/api/send_alerts", methods=["POST"])
@api_login_required
def send_alerts():
    if not SMTP_USER or not SMTP_PASS:
        return jsonify({"success": False,
                        "error": "SMTP_USER / SMTP_PASS not set. Configure them in Render environment variables."})
    try:
        conn = get_db(); c = conn.cursor()
        today = datetime.now().strftime("%Y-%m-%d")
        c.execute("""SELECT s.usn, s.name, s.email FROM students s
                     WHERE s.usn NOT IN (SELECT a.usn FROM attendance a WHERE a.date=?)""", (today,))
        absent = [dict(r) for r in c.fetchall()]; conn.close()

        if not absent:
            return jsonify({"success": True, "message": "No absentees today! Everyone is present.", "sent": 0})

        sent = 0; errors = []

        # Teacher/admin summary
        if ALERT_TO:
            table_rows = "".join(
                f"<tr><td style='padding:8px'>{a['usn']}</td><td style='padding:8px'>{a['name']}</td>"
                f"<td style='padding:8px'>{a['email'] or '—'}</td></tr>" for a in absent)
            summary = f"""
            <div style='font-family:sans-serif;max-width:600px'>
              <h2 style='color:#ef4444'>📋 Absentee Report — {today}</h2>
              <p><strong>{len(absent)}</strong> student(s) absent today:</p>
              <table border='1' cellpadding='0' cellspacing='0'
                style='border-collapse:collapse;width:100%;border-color:#1e2d42'>
                <tr style='background:#0a0e1a;color:#00e5ff'>
                  <th style='padding:10px'>USN</th>
                  <th style='padding:10px'>Name</th>
                  <th style='padding:10px'>Email</th>
                </tr>{table_rows}
              </table>
            </div>"""
            _send_email(ALERT_TO, f"📋 Absentee Report — {today}", summary)
            sent += 1

        # Individual student alerts
        for s in absent:
            if s["email"]:
                body = f"""
                <div style='font-family:sans-serif;max-width:480px'>
                  <h2 style='color:#ef4444'>⚠️ Attendance Alert</h2>
                  <p>Dear <strong>{s['name']}</strong>,</p>
                  <p>You were marked <strong style='color:#ef4444'>absent</strong> on <strong>{today}</strong>.</p>
                  <p>If this is incorrect, please contact your teacher immediately.</p>
                  <br><hr>
                  <p style='color:#64748b;font-size:12px'>AI Smart Attendance System</p>
                </div>"""
                try:
                    _send_email(s["email"], f"⚠️ Attendance Alert — {today}", body)
                    sent += 1
                except Exception as e:
                    errors.append(f"{s['name']}: {e}")

        msg = f"Alerts sent successfully to {sent} recipient(s)."
        if errors: msg += f" Errors: {'; '.join(errors)}"
        return jsonify({"success": True, "message": msg,
                        "absent_count": len(absent), "sent": sent})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

def _send_email(to, subject, html_body):
    msg            = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = ALERT_FROM
    msg["To"]      = to
    msg.attach(MIMEText(html_body, "html"))
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
        srv.starttls()
        srv.login(SMTP_USER, SMTP_PASS)
        srv.sendmail(ALERT_FROM, to, msg.as_string())

# ── ENTRY POINT ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
